from openpyxl import load_workbook
from collections import defaultdict
from gui.program_selection import program_selection

def valueStream():
    wb = load_workbook(filename = "Value Stream 02Apr2025.xlsx")
    posible_programs = wb.sheetnames
    selected_programs:list[str] = program_selection(posible_programs)
    print(selected_programs)





if __name__ == "__main__":
    valueStream()