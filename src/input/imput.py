from openpyxl import load_workbook
from gui.program_selection import program_selection
from datetime import datetime
import easygui
import sys
from copy import copy

def value_stream():
   
    try:
        value_stream_file = load_workbook(filename = "Value Stream MASTER.xlsx", data_only=True)
    except FileNotFoundError as e:
        easygui.msgbox(
            f"The valuestream file could not be found.\n"
            f"Ensure the file is named correctly and in the correct dictionary.\n"
            f"Restart the program\n"
            f"official error: {e}",
            title="Value stream not found"
        )
        sys.exit(1)

    try:
        daily_planning_file = load_workbook(filename=r'/mnt/c/Users/jasper.olthof/Thermo Fisher Scientific/USP Planning - Week Planning/Week Planning.xlsx', data_only=True)
    except FileNotFoundError as e:
        easygui.msgbox(
            f"The daily planning file could not be found.\n"
            f"Ensure the file is named correctly and in the correct dictionary.\n"
            f"Restart the program\n"
            f"official error: {e}",
            title="Daily planning not found"
            )
        sys.exit(1)

    posible_programs:list[str] = value_stream_file.sheetnames

    try:
        selected_programs:list[str] = program_selection(posible_programs)
    except Exception as e:
        easygui.msgbox(
            f"The program selection was unsuccessful.\n"
            f"Restart the program\n"
            f"official error: {e}",
            title="Program selection not performed"
            )
        sys.exit(1)

    if len(selected_programs) == 0:
        easygui.msgbox(
        f"The program selection was unsuccessful.\n"
        f"Restart the program\n",
        title="Program selection not performed"
        )
        sys.exit(1)

    for program in selected_programs:
        batch_number:str = input(f"What is the batch number of {program}? Please put in the full batch number in xx(x).xxx(x) format. ")
        program_sheet = value_stream_file[program]
        replace_batch_number(program_sheet, batch_number)

        for col in program_sheet.iter_cols(min_col=2):
            column_date:datetime = (col[1].value)
            if isinstance(column_date, type(None)):
                break
            if not isinstance(column_date, datetime):
                easygui.msgbox(
                f"The date {column_date} is not formatted correctly.\n"
                f"This error happened during the parsing of {program}.\n"
                "Ensure the correct formatting is used in Excel and that the thaw date is set in the format 01-01-2025 and not 01-Jan-2025.\n"
                "Restart the program.",
                title="Invalid Date Format"
                )
                sys.exit(1)
            column_week_number:int = column_date.isocalendar()[1]
            column_year:int = column_date.isocalendar()[0]
            for cel in col:
                is_coppied = False
                if cel.row <3:
                    continue
                min_row = 0
                max_row = 0
                if not is_top_merged_cell(program_sheet, cel) and is_merged_cell(program_sheet, cel) == True:
                    continue
                if cel.value == None and is_merged_cell(program_sheet, cel) == False:
                    break
                if is_merged_cell(program_sheet, cel) == True:
                    for merged_cell_range in program_sheet.merged_cells.ranges:
                        if cel.coordinate in merged_cell_range:
                            min_row, max_row = merged_cell_range.min_row, merged_cell_range.max_row
                
                try:
                    week_planning = daily_planning_file[f"Week {column_week_number} - {column_year}"]
                except KeyError as e:
                    easygui.msgbox(
                        f"The tab you are trying to use for the daily planning does not exist.\n"
                        f"Please ensure to extend the weekly planning to the end date of the process\n"
                        "Restart the program.",
                        title="Invalid tab in weekly planning"
                    )                    
                    sys.exit(1)

                for coll in week_planning.iter_cols():
                    if is_coppied == True:
                        break
                    if coll[0].value != column_date:
                        continue
                    for cell in coll:
                        if cell.row < 11:
                            continue
                        elif is_merged_cell(week_planning,cell) == True:
                            continue
                        elif cell.value != None:
                            continue
                        else:
                            cell.value = cel.value
                            cell.fill = copy(cel.fill)
                            cell.alignment = copy(cel.alignment)
                            cell.border = copy(cel.border)
                            week_planning.merge_cells(start_row=cell.row, 
                                                        end_row=cell.row + (max_row-min_row),
                                                        start_column=cell.column,
                                                        end_column=cell.column)
                            is_coppied = True
                            
                            break



    reset_batch_number(program_sheet, batch_number)
    value_stream_file.save("Value Stream 02Apr2025.xlsx")
    daily_planning_file.save(r'/mnt/c/Users/jasper.olthof/Thermo Fisher Scientific/USP Planning - Week Planning/Week Planning.xlsx')


def replace_batch_number(program_sheet, batch_number:int) -> None:
    
    for row in program_sheet.iter_rows():
        for cel in row:
            if not isinstance(cel.value, str):
                continue
            data:str = cel.value.strip()
            words:list = data.split(" ")
            new_words = [batch_number if word == "XX.XXX" else word for word in words]
            joined: str = " ".join(new_words)

            if joined != cel.value:
                cel.value = joined

def reset_batch_number(program_sheet, batch_number:int) -> None:

    for row in program_sheet.iter_rows():
        for cel in row:
            if not isinstance(cel.value, str):
                continue
            data:str = cel.value.strip()
            words:list = data.split(" ")
            new_words = ["XX.XXX" if word == batch_number else word for word in words]
            joined: str = " ".join(new_words)

            if joined != cel.value:
                cel.value = joined

def is_merged_cell(program_sheet, cel) -> bool:
    for merged_cell_range in program_sheet.merged_cells.ranges:
        if cel.coordinate in merged_cell_range:
            return True
    return False

def is_top_merged_cell(program_sheet, cel) -> bool:
    for merged_cell_range in program_sheet.merged_cells.ranges:
        if cel.coordinate == merged_cell_range.coord.split(':')[0]:
            return True
    return False

if __name__ == "__main__":
    value_stream()